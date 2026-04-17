"""
app.py — Maltego CSV → Excel local web app
Run: python app.py
Open: http://localhost:5000
"""

import io
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, jsonify, send_file, render_template_string

app = Flask(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
COLORS = {
    'subject':      ('1F3864', 'FFFFFF'),
    'EmailAddress': ('2E4057', 'FFFFFF'),
    'Github':       ('24292E', 'FFFFFF'),
    'Google':       ('4285F4', 'FFFFFF'),
    'Facebook':     ('1877F2', 'FFFFFF'),
    'Duolingo':     ('3A8C00', 'FFFFFF'),
    'Vivino':       ('7B1C1C', 'FFFFFF'),
    'Twitter':      ('1DA1F2', 'FFFFFF'),
    'LinkedIn':     ('0A66C2', 'FFFFFF'),
    'Instagram':    ('C13584', 'FFFFFF'),
    'Affiliation':  ('7B68EE', 'FFFFFF'),
    'Alias':        ('E8F4FD', '000000'),
    'URL':          ('FFF8E7', '555500'),
    'Image':        ('F0F0F0', '666666'),
    'DateTime':     ('FFF0F0', '884444'),
    'Person':       ('E8FFE8', '004400'),
    'PhoneNumber':  ('FFF0FF', '440044'),
    'default':      ('F5F5F5', '333333'),
}

thin = Side(style='thin', color='CCCCCC')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell_style(cell, color_key, bold=False, size=9):
    bg, fg = COLORS.get(color_key, COLORS['default'])
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.font      = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.border    = bdr
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def clean_type(entity_type):
    return entity_type.replace('maltego.affiliation.', '').replace('maltego.', '')


def best_label(row):
    for field in [
        'phonenumber', 'phones.phone-number', 'phone', 'phone-number',
        'email',
        'person.name', 'person.fullname', 'full_name', 'display_name',
        'alias', 'preferred_username', 'name',
        'url', 'title', 'short-title', 'description', 'query',
        'affiliation.uid', 'affiliation.alias',
    ]:
        val = row.get(field)
        if val and str(val) not in ('nan', ''):
            return str(val)
    # EntityType-aware fallback before showing raw EntityID
    etype = str(row.get('EntityType', ''))
    if 'Phone' in etype:
        return f'[phone: {row.name}]'
    return row.name


def find_roots(entities, links):
    """
    Pure structural root detection.
    A node is a root if:
      - It has no incoming edges, OR
      - Every incoming edge comes from a node it also points back to (mutual cycle).
    Affiliation-type nodes (platform accounts) are excluded — they are never roots.
    No entity types are hardcoded as preferred or required.
    """
    all_sources = set(links['SourceEntityID'])

    def is_root(node):
        incoming = set(links[links['TargetEntityID'] == node]['SourceEntityID'])
        if not incoming:
            return True
        outgoing = set(links[links['SourceEntityID'] == node]['TargetEntityID'])
        return incoming.issubset(outgoing)

    def is_affiliation(node):
        if node not in entities.index:
            return False
        return 'affiliation' in entities.loc[node]['EntityType'].lower()

    candidates = [n for n in all_sources if is_root(n) and not is_affiliation(n)]
    if candidates:
        return candidates
    # Last resort: all sources that aren't affiliations
    return [n for n in all_sources if not is_affiliation(n)] or list(entities.index)


def bfs(root_id, links):
    visited, queue, result = set(), [(root_id, None, 0)], []
    while queue:
        node, parent, depth = queue.pop(0)
        if node in visited:
            continue
        visited.add(node)
        result.append((node, parent, depth))
        for child in links[links['SourceEntityID'] == node]['TargetEntityID']:
            if child not in visited:
                queue.append((child, node, depth + 1))
    return result


def build_subject_data(root_id, entities, links):
    nodes   = bfs(root_id, links)
    depth1  = [(n, p) for n, p, d in nodes if d == 1]
    depth2  = [(n, p) for n, p, d in nodes if d == 2]
    root_row   = entities.loc[root_id] if root_id in entities.index else None
    root_label = best_label(root_row) if root_row is not None else root_id
    root_type  = clean_type(root_row['EntityType']) if root_row is not None else 'Entity'
    platforms  = []
    for p1_id, _ in depth1:
        if p1_id not in entities.index:
            continue
        p1_row   = entities.loc[p1_id]
        p1_type  = clean_type(p1_row['EntityType'])
        p1_label = best_label(p1_row)
        children = []
        for c_id, p in depth2:
            if p == p1_id and c_id in entities.index:
                c_row = entities.loc[c_id]
                children.append((clean_type(c_row['EntityType']), best_label(c_row)))
        platforms.append((p1_type, p1_label, children))
    return root_label, root_type, platforms


def process_csv(csv_bytes):
    df       = pd.read_csv(io.BytesIO(csv_bytes))
    entities = df[df['LinkID'].isna()].set_index('EntityID')
    links    = df[df['LinkID'].notna()][['SourceEntityID', 'TargetEntityID']]

    roots        = find_roots(entities, links)
    subject_data = [build_subject_data(r, entities, links) for r in roots]

    max_children = max(
        (len(children) for _, _, plats in subject_data for _, _, children in plats),
        default=0
    )

    seen_platforms = []
    for _, _, plats in subject_data:
        for ptype, _, _ in plats:
            if ptype not in seen_platforms:
                seen_platforms.append(ptype)

    # Build preview JSON
    preview = []
    for root_label, root_type, platforms in subject_data:
        plat_summary = []
        for ptype, plabel, children in platforms:
            plat_summary.append({
                'type': ptype,
                'label': plabel,
                'children': [{'type': ct, 'label': cl} for ct, cl in children],
                'color': '#' + COLORS.get(ptype, COLORS['default'])[0]
            })
        preview.append({
            'label': root_label,
            'type': root_type,
            'platforms': plat_summary
        })

    return subject_data, seen_platforms, max_children, preview


def build_excel_bytes(subject_data, seen_platforms, max_children):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Entity Map'

    col = 1
    ws.cell(1, col, 'Subject');    cell_style(ws.cell(1, col), 'subject', bold=True, size=10);      col += 1
    ws.cell(1, col, 'Root Value'); cell_style(ws.cell(1, col), 'EmailAddress', bold=True, size=10); col += 1

    platform_col_map = {}
    for ptype in seen_platforms:
        platform_col_map[ptype] = col
        ws.cell(1, col, ptype); cell_style(ws.cell(1, col), ptype, bold=True, size=10); col += 1
        for _ in range(max_children):
            ws.cell(1, col, '↳ detail'); cell_style(ws.cell(1, col), ptype, size=9); col += 1

    total_cols = col - 1

    for row_idx, (root_label, root_type, platforms) in enumerate(subject_data, start=2):
        for c in range(1, total_cols + 1):
            ws.cell(row_idx, c, ''); cell_style(ws.cell(row_idx, c), 'default')
        ws.cell(row_idx, 1, root_label); cell_style(ws.cell(row_idx, 1), 'subject', bold=True, size=10)
        ws.cell(row_idx, 2, root_label); cell_style(ws.cell(row_idx, 2), root_type, size=9)
        for ptype, plabel, children in platforms:
            if ptype not in platform_col_map:
                continue
            pc = platform_col_map[ptype]
            ws.cell(row_idx, pc, plabel); cell_style(ws.cell(row_idx, pc), ptype, bold=True, size=9)
            for i, (ctype, clabel) in enumerate(children):
                ws.cell(row_idx, pc + 1 + i, f'{ctype}: {clabel}')
                cell_style(ws.cell(row_idx, pc + 1 + i), ctype, size=8)

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    for c in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24

    ws.row_dimensions[1].height = 28
    for r in range(2, len(subject_data) + 2):
        ws.row_dimensions[r].height = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    with open(os.path.join(os.path.dirname(__file__), 'index.html')) as f:
        return f.read()


@app.route('/preview', methods=['POST'])
def preview():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({'error': 'Please upload a .csv file'}), 400
    try:
        csv_bytes = f.read()
        _, _, _, preview_data = process_csv(csv_bytes)
        # Store csv in app context for download
        app.config['LAST_CSV'] = csv_bytes
        return jsonify({'roots': preview_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download', methods=['POST'])
def download():
    csv_bytes = app.config.get('LAST_CSV')
    if not csv_bytes:
        return jsonify({'error': 'No file loaded — upload first'}), 400
    try:
        subject_data, seen_platforms, max_children, _ = process_csv(csv_bytes)
        buf = build_excel_bytes(subject_data, seen_platforms, max_children)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='maltego_export.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("\n  Maltego → Excel  |  http://localhost:5000\n")
    app.run(debug=True, port=5000)
