"""
app.py — Maltego CSV → Excel local web app
Run: python app.py
Open: http://localhost:5000
"""

import io
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, jsonify, send_file

app = Flask(__name__)

# ── Colour palette ─────────────────────────────────────────────────────────────
COLORS = {
    'subject':                  ('1F3864', 'FFFFFF'),
    'lead_id':                  ('1F3864', 'FFFFFF'),
    'EmailAddress':             ('2E4057', 'FFFFFF'),
    'PhoneNumber':              ('3D2B56', 'FFFFFF'),
    'Person':                   ('0A3D2B', 'FFFFFF'),
    'Github':                   ('24292E', 'FFFFFF'),
    'Google':                   ('4285F4', 'FFFFFF'),
    'Facebook':                 ('1877F2', 'FFFFFF'),
    'Duolingo':                 ('3A8C00', 'FFFFFF'),
    'Vivino':                   ('7B1C1C', 'FFFFFF'),
    'Twitter':                  ('1DA1F2', 'FFFFFF'),
    'LinkedIn':                 ('0A66C2', 'FFFFFF'),
    'Instagram':                ('C13584', 'FFFFFF'),
    'WhatsApp':                 ('25D366', 'FFFFFF'),
    'Telegram':                 ('2CA5E0', 'FFFFFF'),
    'TrueCallerAffiliation':    ('E8A020', 'FFFFFF'),
    'EyeconAffiliation':        ('E05020', 'FFFFFF'),
    'GetContactAffiliation':    ('20A0E0', 'FFFFFF'),
    'Hiya':                     ('6020C0', 'FFFFFF'),
    'Affiliation':              ('7B68EE', 'FFFFFF'),
    'Alias':                    ('E8F4FD', '000000'),
    'URL':                      ('FFF8E7', '555500'),
    'Image':                    ('F0F0F0', '666666'),
    'DateTime':                 ('FFF0F0', '884444'),
    'Location':                 ('E8FFE8', '004400'),
    'default':                  ('F5F5F5', '333333'),
}

thin = Side(style='thin', color='CCCCCC')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell_style(cell, color_key, bold=False, size=9):
    bg, fg = COLORS.get(color_key, COLORS['default'])
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.font      = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.border    = bdr
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def clean_type(t):
    return (t.replace('maltego.affiliation.', '')
             .replace('maltego.', '')
             .replace('test.', ''))


def best_label(row):
    etype = str(row.get('EntityType', '') or '')
    network = str(row.get('affiliation.network', '') or '').strip()

    # Generic maltego.Affiliation: always Profile Exists - <network>
    if etype == 'maltego.Affiliation' and network not in ('', 'nan'):
        return f'Profile Exists - {network}'

    # Platform-specific affiliations (maltego.affiliation.Github etc.):
    # prefer real identifier (alias, uid) — only fall back to Profile Exists if nothing better
    if etype.startswith('maltego.affiliation.'):
        for field in ['affiliation.alias', 'affiliation.uid']:
            val = row.get(field)
            if val and str(val) not in ('nan', ''):
                return str(val).strip()
        if network not in ('', 'nan'):
            return f'Profile Exists - {network}'

    # Standard label lookup
    for field in ['person.fullname', 'email', 'phonenumber', 'name', 'person.name',
                  'display_name', 'alias', 'title', 'surl', 'caller_id',
                  'url', 'datetime', 'location.name', 'description', 'affiliation.uid']:
        val = row.get(field)
        if val and str(val) not in ('nan', ''):
            return str(val).replace('\n', ' | ').strip()
    return str(row.name)


def find_roots(entities, links):
    """
    If the CSV has test.lead_id entities, always use those as roots.
    Otherwise fall back to pure structural detection (no incoming non-cyclic edges),
    excluding affiliation-type nodes.
    """
    lead_entities = entities[entities['EntityType'] == 'test.lead_id']
    if not lead_entities.empty:
        return list(lead_entities.index)

    # Structural fallback
    all_sources = set(links['SourceEntityID'])

    def is_root(node):
        incoming = set(links[links['TargetEntityID'] == node]['SourceEntityID'])
        if not incoming:
            return True
        outgoing = set(links[links['SourceEntityID'] == node]['TargetEntityID'])
        return incoming.issubset(outgoing)

    def is_affiliation(node):
        if node not in entities.index:
            return False
        return 'affiliation' in entities.loc[node]['EntityType'].lower()

    candidates = [n for n in all_sources if is_root(n) and not is_affiliation(n)]
    return candidates if candidates else list(entities.index)


def bfs(root_id, links):
    visited, queue, result = set(), [(root_id, None, 0)], []
    while queue:
        node, parent, depth = queue.pop(0)
        if node in visited:
            continue
        visited.add(node)
        result.append((node, parent, depth))
        for child in links[links['SourceEntityID'] == node]['TargetEntityID']:
            if child not in visited:
                queue.append((child, node, depth + 1))
    return result


def build_row(root_id, entities, links):
    """
    Builds a flat row from a root entity outward.

    Two graph shapes are handled:
      3-level (test.lead_id root):
        lead -> identity(email/phone) -> platform -> leaf
        identity nodes go into identity_nodes, platforms under them into platform_nodes

      2-level (email/phone is root):
        phone/email -> platform -> leaf
        no identity_nodes, depth-1 nodes go directly into platform_nodes
    """
    nodes = bfs(root_id, links)

    IGNORE_TYPES = {'maltego.Image', 'maltego.DateTime'}

    root_row   = entities.loc[root_id] if root_id in entities.index else None
    root_type  = clean_type(root_row['EntityType']) if root_row is not None else 'Entity'
    is_lead    = root_row is not None and root_row['EntityType'] == 'test.lead_id'
    if is_lead and root_row is not None:
        surl = str(root_row.get('surl', '') or '').strip()
        root_label = surl if surl and surl != 'nan' else root_id
    else:
        root_label = best_label(root_row) if root_row is not None else root_id

    identity_nodes = []
    platform_nodes = []

    if is_lead:
        # 3-level: depth-1 are identity nodes (email/phone), depth-2 are platforms
        for n, p, d in nodes:
            if n not in entities.index or n == root_id:
                continue
            row   = entities.loc[n]
            etype = row['EntityType']
            if d == 1:
                identity_nodes.append((clean_type(etype), best_label(row), n))
            elif d == 2:
                leaves = [
                    (clean_type(entities.loc[ln]['EntityType']), best_label(entities.loc[ln]))
                    for ln, lp, ld in nodes
                    if ld == 3 and lp == n
                    and ln in entities.index
                    and entities.loc[ln]['EntityType'] not in IGNORE_TYPES
                ]
                platform_nodes.append((clean_type(etype), best_label(row), p, leaves))
    else:
        # 2-level: depth-1 are platforms, depth-2 are leaves
        for n, p, d in nodes:
            if n not in entities.index or n == root_id:
                continue
            row   = entities.loc[n]
            etype = row['EntityType']
            if d == 1:
                leaves = [
                    (clean_type(entities.loc[ln]['EntityType']), best_label(entities.loc[ln]))
                    for ln, lp, ld in nodes
                    if ld == 2 and lp == n
                    and ln in entities.index
                    and entities.loc[ln]['EntityType'] not in IGNORE_TYPES
                ]
                platform_nodes.append((clean_type(etype), best_label(row), p, leaves))

    return root_label, root_type, identity_nodes, platform_nodes


def process_csv(csv_bytes):
    df       = pd.read_csv(io.BytesIO(csv_bytes))
    entities = df[df['LinkID'].isna()].set_index('EntityID')
    links    = df[df['LinkID'].notna()][['SourceEntityID', 'TargetEntityID']]

    roots    = find_roots(entities, links)
    rows     = [build_row(r, entities, links) for r in roots]

    # Sort platform nodes within each row by leaf count descending
    # so that leaf-rich instances get columns first, and leaf-less instances
    # share no wasted detail columns at the end
    sorted_rows = []
    for root_label, root_type, id_nodes, plat_nodes in rows:
        sorted_plat = sorted(plat_nodes, key=lambda x: len(x[3]), reverse=True)
        sorted_rows.append((root_label, root_type, id_nodes, sorted_plat))
    rows = sorted_rows

    # Collect column schema: identity types, then platform types (in order of appearance)
    seen_identity  = []   # (type, label) — one col per unique identity node across all rows
    seen_platforms = []   # platform types — one col group per type

    for _, _, id_nodes, plat_nodes in rows:
        for itype, ilabel, iid in id_nodes:
            if itype not in seen_identity:
                seen_identity.append(itype)
        for ptype, _, _, _ in plat_nodes:
            if ptype not in seen_platforms:
                seen_platforms.append(ptype)

    from collections import defaultdict as _dd
    _plat_max = _dd(int)
    for _, _, _, plat_nodes in rows:
        for ptype, _, _, leaves in plat_nodes:
            _plat_max[ptype] = max(_plat_max[ptype], len(leaves))
    max_leaves = dict(_plat_max)  # per-platform max leaf count

    # Build preview JSON for the web UI
    preview = []
    for root_label, root_type, id_nodes, plat_nodes in rows:
        platforms = []
        for itype, ilabel, iid in id_nodes:
            platforms.append({'type': itype, 'label': ilabel, 'children': [],
                              'color': '#' + COLORS.get(itype, COLORS['default'])[0]})
        for ptype, plabel, parent_id, leaves in plat_nodes:
            platforms.append({
                'type': ptype, 'label': plabel,
                'children': [{'type': lt, 'label': ll} for lt, ll in leaves],
                'color': '#' + COLORS.get(ptype, COLORS['default'])[0]
            })
        preview.append({'label': root_label, 'type': root_type, 'platforms': platforms})

    return rows, seen_identity, seen_platforms, max_leaves, preview


def build_excel_bytes(rows, seen_identity, seen_platforms, max_leaves):
    from collections import defaultdict
    wb = Workbook()
    ws = wb.active
    ws.title = 'Entity Map'

    has_identity = bool(seen_identity)

    if has_identity:
        # Pre-compute per-instance leaf counts in sorted order (descending) for each (itype, ptype)
        # This gives us the exact column allocation per slot — no empty detail cols
        id_plat_instance_leaves = defaultdict(list)  # (itype, ptype) -> [leaf_count, ...] sorted desc
        for _, _, id_nodes, plat_nodes in rows:
            parent_to_itype = {iid: itype for itype, _, iid in id_nodes}
            row_counts = defaultdict(list)
            for ptype, _, parent_id, leaves in plat_nodes:
                itype = parent_to_itype.get(parent_id, 'unknown')
                row_counts[(itype, ptype)].append(len(leaves))
            for key, counts in row_counts.items():
                # Merge across rows: take the max per slot position
                existing = id_plat_instance_leaves[key]
                for i, c in enumerate(counts):
                    if i < len(existing):
                        existing[i] = max(existing[i], c)
                    else:
                        existing.append(c)

        col = 1
        ws.cell(1, col, 'Lead ID'); cell_style(ws.cell(1, col), 'lead_id', bold=True, size=10); col += 1

        identity_col  = {}
        id_plat_col   = {}    # (itype, ptype) -> [start_col_for_slot_0, start_col_for_slot_1, ...]

        for itype in seen_identity:
            identity_col[itype] = col
            ws.cell(1, col, itype); cell_style(ws.cell(1, col), itype, bold=True, size=10); col += 1
            for ptype in seen_platforms:
                instance_leaves = id_plat_instance_leaves.get((itype, ptype), [])
                if not instance_leaves:
                    continue
                slot_cols = []
                for leaf_count in instance_leaves:
                    slot_cols.append(col)
                    ws.cell(1, col, ptype); cell_style(ws.cell(1, col), ptype, bold=True, size=10); col += 1
                    for _ in range(leaf_count):
                        ws.cell(1, col, '↳ detail'); cell_style(ws.cell(1, col), ptype, size=9); col += 1
                id_plat_col[(itype, ptype)] = slot_cols

        total_cols = col - 1

        for row_idx, (root_label, root_type, id_nodes, plat_nodes) in enumerate(rows, start=2):
            for c in range(1, total_cols + 1):
                ws.cell(row_idx, c, ''); cell_style(ws.cell(row_idx, c), 'default')
            ws.cell(row_idx, 1, root_label); cell_style(ws.cell(row_idx, 1), 'lead_id', bold=True, size=10)
            parent_to_itype = {iid: itype for itype, _, iid in id_nodes}
            for itype, ilabel, _ in id_nodes:
                if itype in identity_col:
                    ws.cell(row_idx, identity_col[itype], ilabel)
                    cell_style(ws.cell(row_idx, identity_col[itype]), itype, size=9)
            written = defaultdict(int)
            for ptype, plabel, parent_id, leaves in plat_nodes:
                itype = parent_to_itype.get(parent_id)
                if not itype or (itype, ptype) not in id_plat_col:
                    continue
                slot      = written[(itype, ptype)]
                slot_cols = id_plat_col[(itype, ptype)]
                if slot >= len(slot_cols):
                    continue
                pc = slot_cols[slot]
                ws.cell(row_idx, pc, plabel); cell_style(ws.cell(row_idx, pc), ptype, bold=True, size=9)
                for i, (ltype, llabel) in enumerate(leaves):
                    ws.cell(row_idx, pc + 1 + i, f'{ltype}: {llabel}')
                    cell_style(ws.cell(row_idx, pc + 1 + i), ltype, size=8)
                written[(itype, ptype)] += 1

    else:
        # 2-level: Root | Plat | details...
        # Same per-instance allocation for 2-level layout
        plat_instance_leaves = defaultdict(list)
        for _, _, _, plat_nodes in rows:
            row_counts = defaultdict(list)
            for ptype, _, _, leaves in plat_nodes:
                row_counts[ptype].append(len(leaves))
            for ptype, counts in row_counts.items():
                existing = plat_instance_leaves[ptype]
                for i, c in enumerate(counts):
                    if i < len(existing):
                        existing[i] = max(existing[i], c)
                    else:
                        existing.append(c)

        col = 1
        ws.cell(1, col, 'Root'); cell_style(ws.cell(1, col), 'lead_id', bold=True, size=10); col += 1

        plat_col = {}
        for ptype in seen_platforms:
            instance_leaves = plat_instance_leaves.get(ptype, [])
            if not instance_leaves:
                continue
            slot_cols = []
            for leaf_count in instance_leaves:
                slot_cols.append(col)
                ws.cell(1, col, ptype); cell_style(ws.cell(1, col), ptype, bold=True, size=10); col += 1
                for _ in range(leaf_count):
                    ws.cell(1, col, '↳ detail'); cell_style(ws.cell(1, col), ptype, size=9); col += 1
            plat_col[ptype] = slot_cols

        total_cols = col - 1

        for row_idx, (root_label, root_type, id_nodes, plat_nodes) in enumerate(rows, start=2):
            for c in range(1, total_cols + 1):
                ws.cell(row_idx, c, ''); cell_style(ws.cell(row_idx, c), 'default')
            ws.cell(row_idx, 1, root_label); cell_style(ws.cell(row_idx, 1), 'lead_id', bold=True, size=10)
            written = defaultdict(int)
            for ptype, plabel, _, leaves in plat_nodes:
                if ptype not in plat_col:
                    continue
                slot      = written[ptype]
                slot_cols = plat_col[ptype]
                if slot >= len(slot_cols):
                    continue
                pc = slot_cols[slot]
                ws.cell(row_idx, pc, plabel); cell_style(ws.cell(row_idx, pc), ptype, bold=True, size=9)
                for i, (ltype, llabel) in enumerate(leaves):
                    ws.cell(row_idx, pc + 1 + i, f'{ltype}: {llabel}')
                    cell_style(ws.cell(row_idx, pc + 1 + i), ltype, size=8)
                written[ptype] += 1

    ws.column_dimensions['A'].width = 20
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24
    ws.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    with open(os.path.join(os.path.dirname(__file__), 'index.html')) as f:
        return f.read()


@app.route('/preview', methods=['POST'])
def preview():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({'error': 'Please upload a .csv file'}), 400
    try:
        csv_bytes = f.read()
        _, _, _, _, preview_data = process_csv(csv_bytes)
        app.config['LAST_CSV'] = csv_bytes
        return jsonify({'roots': preview_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download', methods=['POST'])
def download():
    csv_bytes = app.config.get('LAST_CSV')
    if not csv_bytes:
        return jsonify({'error': 'No file loaded — upload first'}), 400
    try:
        rows, seen_identity, seen_platforms, max_leaves, _ = process_csv(csv_bytes)
        buf = build_excel_bytes(rows, seen_identity, seen_platforms, max_leaves)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='maltego_export.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("\n  Maltego → Excel  |  http://localhost:5000\n")
    app.run(debug=True, port=5000)
